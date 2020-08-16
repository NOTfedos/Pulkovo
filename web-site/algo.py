from datetime import date, timedelta, time
from os import path
import json
from typing import Dict, List, Optional
from vvod_excel import Aud, Program, Teacher, scrap, Data
from catalog import discipline, speciality
from copy import copy
from openpyxl import Workbook

"""      data.json      """


class Lesson:
    def __init__(self, number, course, day, aud, teacher, group):
        self.number = number
        self.course = course
        self.day = day
        self.aud = aud
        self.teacher = teacher
        self.group = group


class Group:
    def __init__(self, prog):
        self.prog = copy(prog)
        self.sch = dict()


def choose_group(name, groups):
    for i, group in enumerate(groups):
        if speciality(group.prog) == name and group.prog.days > 0:
            return i


def choose_teacher(group, teachers):
    for i, teacher in enumerate(teachers):
        if speciality(teacher) == 'Досмотр':
            return i


def choose_auditorium(group, auditoriums):
    for i, auditorium in enumerate(auditoriums):
        if speciality(auditorium) == 'Досмотр':
            return i


def save(done):
    wb = Workbook()
    wsx = wb.active
    now = date(2020, 1, 20)
    for week in range(53):
        ws = wb.create_sheet(f'Week {week + 1}')
        for i in range(7):
            for j in range(4):
                ws.cell(i * 5 + j + 2, 1, now.strftime('%A'))
                ws.cell(i * 5 + j + 2, 2, f'{j + 1} пара')
                for index, group in enumerate(done):
                    lesson = group.sch.get(now)
                    ws.cell(i * 5 + j + 2, 3 + index, lesson.course[j].replace('', '') if lesson is not None else 'Свободно')
            now += timedelta(days=1)
    wb.remove(wsx)
    wb.save(path.join('downloads', 'result.xlsx'))


def process_week(n_week: int, plan: List[Optional[str]], groups: List[Group], teachers, auditoriums, done):
    week = plan[n_week]
    if week is None:
        return False
    splatted_groups = week.split('\n\n')
    for group in splatted_groups:
        left, right = group.split('-')
        start, end = left.split('\n')[-1].strip(), right.split('\n')[0].strip()
        start_date, end_date = date(2020, *map(int, reversed(start.split('.')))), \
                               date(2020, *map(int, reversed(end.split('.'))))
        first = date(2019, 12, 30) + timedelta(days=n_week * 7)
        last = first + timedelta(days=6)
        start_date = max(start_date, first)
        end_date = min(end_date, last)
        name = ' '.join(left.split('\n')[:-1]).strip()
        now = start_date
        while now <= end_date:
            mask = 0
            while True:
                group = choose_group(name, groups[:-mask] if mask > 0 else groups)
                print(group)
                if group is None:
                    break
                teacher = choose_teacher(groups[group], teachers[:-mask] if mask > 0 else teachers)
                if teacher is None:
                    break
                auditorium = choose_auditorium(groups[group], auditoriums[:-mask] if mask > 0 else auditoriums)
                if auditorium is None:
                    break

                if len(groups[group].prog.skelet) == 0:
                    done.append(groups[group])
                    del groups[group]
                else:
                    mask += 1
                    groups[group].sch[now] = Lesson(0, groups[group].prog.skelet[0], now, auditoriums[auditorium],
                                                    teachers[teacher], groups[group])
                    del groups[group].prog.skelet[0]
                    groups = groups[:group] + groups[group:] + [groups[group]]

                teachers = teachers[:teacher] + teachers[teacher:] + [teachers[teacher]]
            now += timedelta(days=1)
    return True


def proc():
    data = scrap()
    plan = data.disp['Досмотр']
    programs = list()
    for program in data.progs:
        if discipline(program) == 'Досмотр' and program.skelet != 0:
            programs.append(program)
    teachers = list()
    for teacher in data.teachers:
        if discipline(teacher) == 'Досмотр':
            teachers.append(teacher)
    auditoriums = list()
    for auditorium in data.audit:
        if discipline(auditorium) == 'Досмотр':
            auditoriums.append(auditorium)
    groups = list()
    for program in programs:
        for i in range(int(program.group)):
            groups.append(Group(program))
    done = list()
    week = process_week(4, plan, groups, teachers, auditoriums, done)
    print('Done')
    save(done)
    return done


'''
    current_day = date(2019, 12, 30)

    group_schedule = dict()
    teacher_schedule = dict()
    data = scrap()
    plan: Dict[str, List] = data.disp  # - словарь массивов (приложение 1)
    auditoriums: List[Aud] = data.audit  # - массив классов аудиторий (приложение 2)
    programs: List[Program] = data.progs  # - массив классов программ (приложение 2)
    teachers: List[Teacher] = data.teachers  # - массив классов учителей (приложение 2)
    for program in programs:  # формируем словарь расписаний групп
        if program.group is None or program.skelet == 0:
            print(f'No data for {program.name}')
            continue
        for i in range(int(program.group)):
            group_schedule.update({Group(program, priority=i): dict()})

    for teacher in teachers:  # формируем словарь расписаний учителей
        teacher_schedule.update({teacher: dict()})

    process_week(data, 3)

    # ----------------------------- КОНЕЦ ИНИЦИАЛИЗАЦИИ ---------------------------------

    while current_day.year < 2021:  # проход по каждому дню
        for i in range(4):
            if i == 0:
                for group in group_schedule.keys():  # добавляем текущий день в расписание группы
                    group_schedule[group].update({current_day: []})
                for teacher in teacher_schedule.keys():
                    teacher_schedule[teacher].update({current_day: []})
            for group in group_schedule.keys():
                plan[disciplines_indexer[group.prog.disp]]

                if program.index <= len(program.skelet):  # если группа ещё не прошла программу
                    # ищем преподавателей на эту программу
                    program.index += 1

        current_day += timedelta(days=1)
'''

if __name__ == '__main__':
    proc()
