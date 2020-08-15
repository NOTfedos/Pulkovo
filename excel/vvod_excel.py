import openpyxl
import os
import win32com.client as win32


# ----------------------------------------------------------
class Aud(object):
    def __init__(self, num, space, spes, priority, whitelist, typ):
        self.num = num
        self.space = space
        self.spes = spes  # особенности
        self.priority = priority
        self.whitelist = whitelist
        self.type = typ


class Program(object):
    def __init__(self, num, disp, name, spes, time, days, group, people, people_vn, teach, aud, skelet):
        self.num = num
        self.disp = disp
        self.name = name
        self.spes = spes  # особенности
        self.time = time
        self.days = days
        self.group = group
        self.people = people
        self.people_vn = people_vn
        self.teach = teach
        self.aud = aud
        self.skelet = skelet
        self.index = 1


class Teacher(object):
    def __init__(self, num, name, disp, programs, priority, blacklist, graph, smen):
        self.num = num
        self.name = name
        self.disp = disp
        self.programs = programs
        self.priority = priority
        self.blacklist = blacklist
        self.graph = graph
        self.smen = smen


def whitelist_check(s):
    for c, i in enumerate(s):
        if "кроме" in i:
            s.pop(c)
    return s


def pars_lessons(ind, file_1):
    return list(map(lambda s: s.value, file_1['I' + str(ind):'DK' + str(ind)][0][::2]))


# ----------------------------------------------------------
file_1 = openpyxl.load_workbook('Приложение №1.xlsx')['ДПО']

disp = {}

offset_1 = 6

index = 6

while file_1['G' + str(index)].value is not None:
    name = file_1['G' + str(index)].value
    disp[name] = pars_lessons(index, file_1)
    index += 2

del file_1

# ----------------------------------------------------------
file_2 = openpyxl.load_workbook('Приложение №2.xlsx')
# ---------------
file_auds = file_2['параметры аудиторий']

Audit = []
index = 2

while file_auds['A' + str(index)].value is not None:
    if 'теоретические' in file_auds['C' + str(index)].value:
        Audit.append(Aud(file_auds['A' + str(index)].value, file_auds['B' + str(index)].value,
                         file_auds['D' + str(index)].value.split(', '), file_auds['E' + str(index)].value,
                         whitelist_check(file_auds['F' + str(index)].value.split('; ').copy()),
                         file_auds['C' + str(index)].value))
    index += 1
del file_auds
# ---------------

file_programs = file_2['параметры программ']

index = 2
Programs = []
last = ''

while file_programs['A' + str(index)].value is not None:
    if (last != file_programs['B' + str(index)].value) and (file_programs['B' + str(index)].value is not None):
        last = file_programs['B' + str(index)].value
    if (str(file_programs['C' + str(index)].value) + '.doc') in os.listdir('Приложение №3 (2)'):
        word = win32.Dispatch("Word.Application")
        word.Visible = 0
        word.Documents.Open("\\Приложение №3 (2)\\" + str(file_programs['C' + str(index)].value) + '.doc')
        table = word.ActiveDocument.Tables(2)
        skelet = []
        for i in range((table.Rows.Count - 1) // 4):
            skelet.append(
                [table.Cell(Row=i * 4 + 1, Column=3).Range.Text, table.Cell(Row=i * 4 + 2, Column=3).Range.Text,
                 table.Cell(Row=i * 4 + 3, Column=3).Range.Text, table.Cell(Row=i * 4 + 4, Column=3).Range.Text])

    else:
        skelet = 0

    Programs.append(Program(file_programs['A' + str(index)].value, last, file_programs['C' + str(index)].value,
                            file_programs['D' + str(index)].value.split(', '), file_programs['F' + str(index)].value,
                            file_programs['I' + str(index)].value, file_programs['J' + str(index)].value,
                            file_programs['K' + str(index)].value, file_programs['L' + str(index)].value,
                            file_programs['M' + str(index)].value, file_programs['N' + str(index)].value, skelet))
    index += 1

del file_programs
# ---------------
file_4 = openpyxl.load_workbook('Приложение №4.xlsx')['график смен']

index = 4
smen = [[] for _ in range(4)]

for i in range(12):
    smen[int(file_4['A' + str(i * 7 + 5)].value[-1]) - 1].append(
        list(map(lambda s: s.value, file_4['B' + str(index):'AF' + str(index)][0][::2])))
    smen[int(file_4['A' + str(i * 7 + 6)].value[-1]) - 1].append(
        list(map(lambda s: s.value, file_4['B' + str(index):'AF' + str(index)][0][::2])))
    smen[int(file_4['A' + str(i * 7 + 7)].value[-1]) - 1].append(
        list(map(lambda s: s.value, file_4['B' + str(index):'AF' + str(index)][0][::2])))
    smen[int(file_4['A' + str(i * 7 + 8)].value[-1]) - 1].append(
        list(map(lambda s: s.value, file_4['B' + str(index):'AF' + str(index)][0][::2])))

del file_4

# ---------------
file_5 = openpyxl.load_workbook('Приложение №5.xlsx')['План']

index = 1
while (file_5['A' + str(index)].value == None):
    index += 1

otpusk = []

while (file_5['A' + str(index)].value.isnumeric() == None):
    otpusk.append([file_5['B' + str(index)].value,
                   list(map(lambda s: s.value, file_1['D' + str(index):'AA' + str(index)][0][::2]))])
    index += 1

del file_5

# ---------------
file_teachers = file_2['параметры преподавателей']

index = 2
Teachers = []

while (file_teachers['A' + str(index)].value != None):
    Teachers.append(Teacher(file_teachers['A' + str(index)].value, file_teachers['b' + str(index)].value,
                            file_teachers['C' + str(index)].value,
                            str(file_teachers['D' + str(index)].value).split(';'),
                            file_teachers['E' + str(index)].value, file_teachers['F' + str(index)].value,
                            file_teachers['G' + str(index)].value, file_teachers['H' + str(index)].value))
    index += 1
